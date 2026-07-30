import io
import numpy as np
import openpyxl
import pandas as pd
import plotly.express as px
import streamlit as st

# (ReportLab imports removed)

# Set Streamlit Page Configuration
st.set_page_config(
    page_title="Gym Performance Dashboard", page_icon="🏋️‍♂️", layout="wide"
)


# -----------------------------------------------------------------------------
# 1. DATA LOADING & PREPROCESSING
# -----------------------------------------------------------------------------
@st.cache_data
def load_data():
  file_path = "Gym_Registre_PG.xlsx"
  df = pd.read_excel(file_path, sheet_name="Registre_Exercicis", header=3)

  df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
  df["Pes (kg)"] = pd.to_numeric(df["Pes (kg)"], errors="coerce").fillna(0)
  df["Repeticions"] = pd.to_numeric(df["Repeticions"], errors="coerce").fillna(
      0
  )
  df["Temps (min)"] = pd.to_numeric(df["Temps (min)"], errors="coerce").fillna(
      0
  )

  df = df.dropna(subset=["Exercici", "Data"]).copy()
  df["Exercici"] = df["Exercici"].astype(str).str.strip()
  df["Grup Muscular"] = df["Grup Muscular"].astype(str).str.strip()

  # Set Volume & Estimated 1RM (Brzycki Formula: Weight * 36 / (37 - Reps))
  df["Set_Volume"] = df["Pes (kg)"] * df["Repeticions"]
  df["Estimated_1RM"] = df.apply(
      lambda r: r["Pes (kg)"] * (36 / (37 - r["Repeticions"]))
      if (r["Pes (kg)"] > 0 and 0 < r["Repeticions"] < 37)
      else 0,
      axis=1,
  )
  return df


# -----------------------------------------------------------------------------
# HELPER: PERSISTENT EXCEL LOGGING
# -----------------------------------------------------------------------------
def append_set_to_excel(new_data):
  """Appends a new set entry adhering to the exact column layout in Excel."""
  excel_path = "Gym_Registre_PG.xlsx"

  wb = openpyxl.load_workbook(excel_path)
  ws = (
      wb["Registre_Exercicis"]
      if "Registre_Exercicis" in wb.sheetnames
      else wb.active
  )

  # Find the true first empty row in Column A
  target_row = 5
  set_count = 1

  while ws.cell(row=target_row, column=1).value is not None:
    # Auto-calculate the next 'Sèrie' number for the same exercise on the same date
    row_date = ws.cell(row=target_row, column=1).value
    row_ex = ws.cell(row=target_row, column=3).value

    # Convert cell date to DD/MM/YYYY string for comparison
    if isinstance(row_date, pd.Timestamp) or hasattr(row_date, "strftime"):
      row_date_str = row_date.strftime("%d/%m/%Y")
    else:
      row_date_str = str(row_date)

    if (
        row_date_str == new_data["Data"]
        and str(row_ex).strip() == new_data["Exercici"]
    ):
      set_count += 1

    target_row += 1

  # Calculate volume (Weight * Reps)
  volum = new_data["Pes (kg)"] * new_data["Repeticions"]

  # Map directly to Excel columns:
  # Col 1: Data (DD/MM/YYYY)
  # Col 2: Grup Muscular
  # Col 3: Exercici
  # Col 4: Sèrie
  # Col 5: Pes (kg)
  # Col 6: Repeticions
  # Col 7: Volum (kg)
  # Col 8: Temps (min)
  ws.cell(row=target_row, column=1, value=new_data["Data"])
  ws.cell(row=target_row, column=2, value=new_data["Grup Muscular"])
  ws.cell(row=target_row, column=3, value=new_data["Exercici"])
  ws.cell(row=target_row, column=4, value=set_count)
  ws.cell(
      row=target_row,
      column=5,
      value=new_data["Pes (kg)"] if new_data["Pes (kg)"] > 0 else None,
  )
  ws.cell(
      row=target_row,
      column=6,
      value=new_data["Repeticions"] if new_data["Repeticions"] > 0 else None,
  )
  ws.cell(row=target_row, column=7, value=volum)
  ws.cell(
      row=target_row,
      column=8,
      value=new_data["Temps (min)"] if new_data["Temps (min)"] > 0 else None,
  )

  wb.save(excel_path)
  st.cache_data.clear()

# (Section 2: PDF Logic has been completely removed)

# -----------------------------------------------------------------------------
# 3. SIDEBAR CONTROLS, FILTERS & NEW SET ENTRY FORM
# -----------------------------------------------------------------------------
st.sidebar.title("🏋️ Dashboard Controls")

# Always load the persistent local file
df = load_data()

# --- 3A. ENTRY FORM FOR LOGGING NEW SETS DIRECTLY TO EXCEL ---
st.sidebar.markdown("---")

# Use the expander container directly
with st.sidebar.expander("➕ **Log New Set to Excel**", expanded=True):

  # 1. Muscle Group selection INSIDE the expander (not st.sidebar.selectbox)
  existing_mg = sorted(df["Grup Muscular"].unique())
  selected_log_mg = st.selectbox(
      "Muscle Group", options=existing_mg, key="form_mg_select"
  )

  # 2. Filter available exercises dynamically based on selected Muscle Group
  existing_ex = sorted(
      df[df["Grup Muscular"] == selected_log_mg]["Exercici"].unique()
  )

  # 3. Form placed directly inside the expander
  with st.form("log_set_form", clear_on_submit=True):
    log_date = st.date_input("Date", value=pd.Timestamp.now().date())

    log_ex_option = st.selectbox(
        "Exercise", options=existing_ex + ["+ Add New Exercise..."]
    )

    if log_ex_option == "+ Add New Exercise...":
      log_ex = st.text_input("Enter New Exercise Name")
    else:
      log_ex = log_ex_option

    col_weight, col_reps = st.columns(2)
    with col_weight:
      log_weight = st.number_input(
          "Weight (kg)",
          min_value=0.0,
          step=0.5,
          value=None,
          placeholder="0.0",
      )
    with col_reps:
      log_reps = st.number_input(
          "Reps", min_value=0, step=1, value=None, format="%d", placeholder="0"
      )

    log_time = st.number_input(
        "Duration (min)",
        min_value=0.0,
        step=0.5,
        value=None,
        placeholder="0.0",
    )

    submit_set = st.form_submit_button("💾 Save Set to Excel")

    if submit_set:
      if not log_ex.strip():
        st.error("Please enter a valid exercise name.")
      else:
        new_entry = {
            "Data": pd.to_datetime(log_date).strftime("%d/%m/%Y"),
            "Exercici": log_ex.strip(),
            "Grup Muscular": selected_log_mg.strip(),
            "Pes (kg)": log_weight if log_weight is not None else 0.0,
            "Repeticions": log_reps if log_reps is not None else 0,
            "Temps (min)": log_time if log_time is not None else 0.0,
        }

        try:
          append_set_to_excel(new_entry)
          st.success(f"Saved: {log_ex} ({log_weight}kg x {log_reps} reps)!")
          st.rerun()
        except Exception as e:
          st.error(f"Error saving set: {e}")

# -----------------------------------------------------------------------------
# GLOBAL FILTERS (EXPANDER)
# -----------------------------------------------------------------------------
with st.sidebar.expander("🔍 Filters", expanded=True):
  # 1. Date Range Filter
  min_date = df["Data"].min().date()
  max_date = df["Data"].max().date()

  date_range = st.date_input(
      "Select Date Range",
      value=(min_date, max_date),
      min_value=min_date,
      max_value=max_date,
  )

  # Handle single date selection vs full range safely
  if isinstance(date_range, tuple) and len(date_range) == 2:
    start_date, end_date = date_range
  elif isinstance(date_range, tuple) and len(date_range) == 1:
    start_date = end_date = date_range[0]
  else:
    start_date = end_date = date_range

  # 2. Muscle Group Filter
  all_mg = sorted(df["Grup Muscular"].unique().tolist())
  selected_mg = st.multiselect(
      "Muscle Groups", options=all_mg, default=all_mg
  )

  # Dynamic exercise options based on selected muscle groups
  available_exercises = sorted(
      df[df["Grup Muscular"].isin(selected_mg)]["Exercici"].unique().tolist()
  )
  selected_exercises = st.multiselect(
      "Exercises", options=available_exercises, default=available_exercises
  )

# Apply global filtering for dashboard views
df_filtered = df[
    (df["Data"].dt.date >= start_date)
    & (df["Data"].dt.date <= end_date)
    & (df["Grup Muscular"].isin(selected_mg))
    & (df["Exercici"].isin(selected_exercises))
]

# -----------------------------------------------------------------------------
# 4. MAIN DASHBOARD UI
# -----------------------------------------------------------------------------
st.title("🏋️‍♂️ Interactive Gym Performance Dashboard")
st.markdown(
    "Track set volume, e1RM estimations, duration, reps, and personal records"
    " interactively."
)

if df_filtered.empty:
  st.warning(
      "No data available for the selected filters. Please expand your"
      " selections in the sidebar."
  )
  st.stop()

st.markdown("---")

tab1, tab2, tab3 = st.tabs([
    "🏆 Personal Records (Max Day)",
    "📈 Progress & Trends",
    "📊 Muscle Distribution & Raw Data",
])

# =============================================================================
# TAB 1: PERSONAL RECORDS & PER-EXERCISE SUMMARY
# =============================================================================
with tab1:
  st.subheader("🏆 Dia de Màxim Rendiment per Exercici")
  st.caption(
      "Resum del dia de màxim registre (Volum, Temps o Repeticions) per a cada"
      " exercici seleccionat."
  )

  def format_set(row):
    p = row["Pes (kg)"]
    r = row["Repeticions"]
    t = row["Temps (min)"]

    pes_str = f"{int(p)}" if p == int(p) else f"{p}"
    reps_str = f"{int(r)}" if r == int(r) else f"{r}"
    temps_str = f"{int(t)}" if t == int(t) else f"{t}"

    if t > 0:
      return (
          f"{temps_str} min"
          if p == 0 and r == 0
          else f"{pes_str}kg x {reps_str} reps ({temps_str} min)"
      )
    return f"{pes_str}kg x {reps_str} reps" if p > 0 else f"{reps_str} reps"

  df_tab1 = df_filtered.copy()
  df_tab1["Set_Desc"] = df_tab1.apply(format_set, axis=1)

  df_tab1["Ex_Type"] = np.where(
      (df_tab1["Temps (min)"] > 0) & (df_tab1["Set_Volume"] == 0),
      "Timed",
      np.where(
          (df_tab1["Pes (kg)"] == 0)
          & (df_tab1["Temps (min)"] == 0)
          & (df_tab1["Repeticions"] > 0),
          "BW_Reps",
          "Weighted",
      ),
  )

  daily_totals = (
      df_tab1.groupby(["Exercici", "Grup Muscular", "Ex_Type", "Data"])
      .agg(
          Daily_Volume=("Set_Volume", "sum"),
          Daily_Time=("Temps (min)", "sum"),
          Daily_Reps=("Repeticions", "sum"),
      )
      .reset_index()
  )

  daily_totals["Max_Metric_Value"] = np.where(
      daily_totals["Ex_Type"] == "Timed",
      daily_totals["Daily_Time"],
      np.where(
          daily_totals["Ex_Type"] == "BW_Reps",
          daily_totals["Daily_Reps"],
          daily_totals["Daily_Volume"],
      ),
  )

  max_dates = daily_totals.loc[
      daily_totals.groupby("Exercici")["Max_Metric_Value"].idxmax()
  ]

  max_day_sets = pd.merge(
      df_tab1,
      max_dates[["Exercici", "Data", "Max_Metric_Value"]],
      on=["Exercici", "Data"],
      how="inner",
  )

  max_summary = (
      max_day_sets.groupby(
          ["Exercici", "Grup Muscular", "Ex_Type", "Data", "Max_Metric_Value"]
      )
      .apply(
          lambda g: "<br>".join([
              f"<b>Sèrie {i+1}:</b> {row['Set_Desc']}"
              for i, (_, row) in enumerate(g.iterrows())
          ])
      )
      .reset_index(name="Detalls")
  )

  def format_max_record(row):
    val = row["Max_Metric_Value"]
    ex_type = row["Ex_Type"]
    if ex_type == "Timed":
      val_str = f"{int(val)}" if val == int(val) else f"{val:.1f}"
      return f"{val_str} min"
    elif ex_type == "BW_Reps":
      return f"{int(val):,} reps"
    else:
      return f"{val:,.0f} kg"

  max_summary["Registre Màxim"] = max_summary.apply(
      format_max_record, axis=1
  )
  max_summary = max_summary.sort_values(
      by="Max_Metric_Value", ascending=False
  ).reset_index(drop=True)
  max_summary["Data del Màxim"] = max_summary["Data"].dt.strftime("%d/%m/%Y")

  table_df = max_summary[[
      "Exercici",
      "Grup Muscular",
      "Data del Màxim",
      "Registre Màxim",
      "Detalls",
  ]]

  st.markdown(
      table_df.to_html(escape=False, index=False), unsafe_allow_html=True
  )

  st.markdown("---")

  st.subheader("📊 Resum Totals per Exercici")
  ex_summary = (
      df_filtered.groupby(["Exercici", "Grup Muscular"])
      .agg(
          Sessions=("Data", "nunique"),
          Total_Volume=("Set_Volume", "sum"),
          Total_Time=("Temps (min)", "sum"),
          Total_Sets=("Set_Volume", "count"),
          Total_Reps=("Repeticions", "sum"),
      )
      .reset_index()
  )

  ex_summary = ex_summary.sort_values(
      by="Total_Volume", ascending=False
  ).reset_index(drop=True)

  ex_summary["Tonatge Total"] = ex_summary["Total_Volume"].apply(
      lambda x: f"{x:,.0f} kg" if x > 0 else "-"
  )
  ex_summary["Temps Total"] = ex_summary["Total_Time"].apply(
      lambda x: f"{x:,.1f} min" if x > 0 else "-"
  )
  ex_summary["Repeticions Totals"] = ex_summary["Total_Reps"].apply(
      lambda x: f"{int(x):,}" if x > 0 else "-"
  )

  st.dataframe(
      ex_summary[[
          "Exercici",
          "Grup Muscular",
          "Sessions",
          "Tonatge Total",
          "Temps Total",
          "Total_Sets",
          "Repeticions Totals",
      ]].rename(
          columns={
              "Sessions": "Sessions Totals",
              "Total_Sets": "Sèries Totals",
          }
      ),
      use_container_width=True,
      hide_index=True,
  )

# =============================================================================
# TAB 2: PROGRESSION OVER TIME
# =============================================================================
with tab2:
  df_filtered_copy = df_filtered.copy()
  df_filtered_copy["Set_Desc"] = df_filtered_copy.apply(format_set, axis=1)

  is_timed_ex = (df_filtered_copy["Temps (min)"] > 0) & (
      df_filtered_copy["Set_Volume"] == 0
  )
  is_bodyweight_reps_ex = (
      (df_filtered_copy["Pes (kg)"] == 0)
      & (df_filtered_copy["Temps (min)"] == 0)
      & (df_filtered_copy["Repeticions"] > 0)
  )
  is_weighted_ex = ~is_timed_ex & ~is_bodyweight_reps_ex

  df_weighted = df_filtered_copy[is_weighted_ex]
  df_timed = df_filtered_copy[is_timed_ex]
  df_bw_reps = df_filtered_copy[is_bodyweight_reps_ex]

  # --- GRAPH 1A: VOLUME ---
  st.subheader("📈 Daily Volume Progression Over Time")
  if not df_weighted.empty:
    daily_vol_summary = (
        df_weighted.groupby(["Exercici", "Grup Muscular", "Data"])
        .apply(
            lambda g: pd.Series({
                "Daily_Volume": g["Set_Volume"].sum(),
                "Max_1RM": g["Estimated_1RM"].max(),
                "Set_Details": "<br>".join([
                    f"Sèrie {i+1}: {row['Set_Desc']}"
                    for i, (_, row) in enumerate(g.iterrows())
                ]),
            })
        )
        .reset_index()
    )

    fig_vol = px.line(
        daily_vol_summary,
        x="Data",
        y="Daily_Volume",
        color="Exercici",
        markers=True,
        labels={"Daily_Volume": "Total Volume (kg)", "Data": "Date"},
        title="Interactive Daily Volume (kg)",
        custom_data=["Exercici", "Daily_Volume", "Set_Details"],
    )
    fig_vol.update_traces(
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "Data: %{x|%d/%m/%Y}<br>"
            "Volum Total: %{customdata[1]:,.0f} kg<br><br>"
            "<b>Detall de Sèries:</b><br>"
            "%{customdata[2]}<extra></extra>"
        )
    )
    fig_vol.update_layout(
        template="plotly_white", height=450, hovermode="closest"
    )
    st.plotly_chart(fig_vol, use_container_width=True)
  else:
    st.info("No volume-based exercise data available for current selection.")

  # --- GRAPH 1B: REPETITIONS ---
  if not df_bw_reps.empty:
    st.markdown("---")
    st.subheader("🔢 Daily Total Repetitions Progression Over Time")

    daily_reps_summary = (
        df_bw_reps.groupby(["Exercici", "Grup Muscular", "Data"])
        .apply(
            lambda g: pd.Series({
                "Daily_Reps": g["Repeticions"].sum(),
                "Set_Details": "<br>".join([
                    f"Sèrie {i+1}: {row['Set_Desc']}"
                    for i, (_, row) in enumerate(g.iterrows())
                ]),
            })
        )
        .reset_index()
    )

    fig_reps = px.line(
        daily_reps_summary,
        x="Data",
        y="Daily_Reps",
        color="Exercici",
        markers=True,
        labels={"Daily_Reps": "Total Repetitions", "Data": "Date"},
        title="Interactive Daily Total Repetitions",
        custom_data=["Exercici", "Daily_Reps", "Set_Details"],
    )
    fig_reps.update_traces(
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "Data: %{x|%d/%m/%Y}<br>"
            "Repeticions Totals: %{customdata[1]:,.0f}<br><br>"
            "<b>Detall de Sèries:</b><br>"
            "%{customdata[2]}<extra></extra>"
        )
    )
    fig_reps.update_layout(
        template="plotly_white", height=450, hovermode="closest"
    )
    st.plotly_chart(fig_reps, use_container_width=True)

  # --- GRAPH 1C: DURATION ---
  if not df_timed.empty:
    st.markdown("---")
    st.subheader("⏱️ Daily Duration Progression Over Time")

    daily_time_summary = (
        df_timed.groupby(["Exercici", "Grup Muscular", "Data"])
        .apply(
            lambda g: pd.Series({
                "Daily_Time": g["Temps (min)"].sum(),
                "Set_Details": "<br>".join([
                    f"Sèrie {i+1}: {row['Set_Desc']}"
                    for i, (_, row) in enumerate(g.iterrows())
                ]),
            })
        )
        .reset_index()
    )

    fig_time = px.line(
        daily_time_summary,
        x="Data",
        y="Daily_Time",
        color="Exercici",
        markers=True,
        labels={"Daily_Time": "Total Duration (min)", "Data": "Date"},
        title="Interactive Daily Duration (minutes)",
        custom_data=["Exercici", "Daily_Time", "Set_Details"],
    )
    fig_time.update_traces(
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            "Data: %{x|%d/%m/%Y}<br>"
            "Temps Total: %{customdata[1]:,.1f} min<br><br>"
            "<b>Detall de Sèries:</b><br>"
            "%{customdata[2]}<extra></extra>"
        )
    )
    fig_time.update_layout(
        template="plotly_white", height=450, hovermode="closest"
    )
    st.plotly_chart(fig_time, use_container_width=True)

  st.markdown("---")

  # --- GRAPH 2: ESTIMATED 1RM ---
  st.subheader("🔥 Estimated 1RM (Brzycki) Progression Over Time")
  if not df_weighted.empty and "daily_vol_summary" in locals():
    daily_1rm = daily_vol_summary[daily_vol_summary["Max_1RM"] > 0]

    if not daily_1rm.empty:
      fig_1rm = px.line(
          daily_1rm,
          x="Data",
          y="Max_1RM",
          color="Exercici",
          markers=True,
          labels={"Max_1RM": "Estimated 1RM (kg)", "Data": "Date"},
          title="Interactive Daily Estimated 1RM (kg)",
          custom_data=["Exercici", "Max_1RM", "Set_Details"],
      )
      fig_1rm.update_traces(
          hovertemplate=(
              "<b>%{customdata[0]}</b><br>"
              "Data: %{x|%d/%m/%Y}<br>"
              "e1RM Màxim: %{customdata[1]:,.1f} kg<br><br>"
              "<b>Detall de Sèries:</b><br>"
              "%{customdata[2]}<extra></extra>"
          )
      )
      fig_1rm.update_layout(
          template="plotly_white", height=450, hovermode="closest"
      )
      st.plotly_chart(fig_1rm, use_container_width=True)
    else:
      st.info("No weight-bearing 1RM data available for selected exercises.")
  else:
    st.info("No weight-bearing 1RM data available for selected exercises.")

import calendar
import streamlit.components.v1 as components

# =============================================================================
# TAB 3: CALENDAR VIEW, MUSCLE DISTRIBUTION & RAW DATA
# =============================================================================
with tab3:
  # -------------------------------------------------------------------------
  # 1. MONTHLY CALENDAR VIEW
  # -------------------------------------------------------------------------
  st.subheader("📅 Gym Activity Calendar")
  st.caption(
      "Overview of exercises performed on each day, filtered by your sidebar"
      " selection."
  )

  # Dynamic Color Palette (10 distinct high-contrast colors)
  color_palette = [
      "#2563EB",  # Blue
      "#059669",  # Emerald
      "#D97706",  # Amber
      "#DC2626",  # Red
      "#7C3AED",  # Violet
      "#DB2777",  # Pink
      "#0D9488",  # Teal
      "#4F46E5",  # Indigo
      "#EA580C",  # Orange
      "#0891B2",  # Cyan
  ]

  # Map each unique exercise in the dataset to a specific color class
  unique_exercises = sorted(df["Exercici"].unique().tolist())
  ex_class_map = {
      ex: f"ex-color-{i % len(color_palette)}"
      for i, ex in enumerate(unique_exercises)
  }

  # Build dynamic CSS rules for each color
  dynamic_color_css = "\n".join([
      f".ex-color-{i} {{ background-color: {color} !important; }}"
      for i, color in enumerate(color_palette)
  ])

  # Select Month and Year based on overall data available
  cal_col1, cal_col2 = st.columns(2)
  with cal_col1:
    selected_year = st.selectbox(
        "Year",
        options=sorted(df["Data"].dt.year.unique(), reverse=True),
        key="cal_year",
    )
  with cal_col2:
    months_map = {
        1: "January",
        2: "February",
        3: "March",
        4: "April",
        5: "May",
        6: "June",
        7: "July",
        8: "August",
        9: "September",
        10: "October",
        11: "November",
        12: "December",
    }
    selected_month_num = st.selectbox(
        "Month",
        options=list(months_map.keys()),
        format_func=lambda x: months_map[x],
        index=pd.Timestamp.now().month - 1,
        key="cal_month",
    )

  # Filter USING df_filtered (respecting Muscle Group and Exercise sidebar selections)
  df_month = df_filtered[
      (df_filtered["Data"].dt.year == selected_year)
      & (df_filtered["Data"].dt.month == selected_month_num)
  ]

  # Group exercises by day
  daily_exercises = (
      df_month.groupby(df_month["Data"].dt.day)["Exercici"]
      .unique()
      .apply(list)
      .to_dict()
  )

  # Calendar layout computation
  month_cal = calendar.monthcalendar(selected_year, selected_month_num)
  days_of_week = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

  # Assemble complete HTML page inside a string
  full_html_doc = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
            margin: 0;
            padding: 0;
            background-color: transparent;
        }}
        .cal-table {{ 
            width: 100%; 
            border-collapse: collapse; 
            table-layout: fixed;
        }}
        .cal-th {{ 
            background-color: #1E293B; 
            color: white; 
            text-align: center; 
            padding: 8px; 
            border: 1px solid #334155; 
            font-size: 13px; 
        }}
        .cal-td {{ 
            vertical-align: top; 
            height: 100px; 
            border: 1px solid #E2E8F0; 
            padding: 4px; 
            background-color: #F8FAFC; 
            overflow-y: auto;
        }}
        .cal-empty {{ 
            background-color: #F1F5F9; 
            border: 1px solid #E2E8F0; 
        }}
        .cal-day-num {{ 
            font-weight: bold; 
            font-size: 11px; 
            color: #475569; 
            margin-bottom: 4px; 
        }}
        .cal-badge {{ 
            color: white; 
            padding: 2px 5px; 
            border-radius: 4px; 
            font-size: 10px; 
            font-weight: 600; 
            margin-bottom: 2px; 
            display: block; 
            white-space: nowrap; 
            overflow: hidden; 
            text-overflow: ellipsis; 
        }}
        .cal-today {{ 
            background-color: #EFF6FF; 
            border: 2px solid #2563EB; 
        }}
        {dynamic_color_css}
    </style>
    </head>
    <body>
    <table class="cal-table">
        <thead>
            <tr>
    """

  for day in days_of_week:
    full_html_doc += f'<th class="cal-th">{day}</th>'
  full_html_doc += "</tr></thead><tbody>"

  today = pd.Timestamp.now().date()

  for week in month_cal:
    full_html_doc += "<tr>"
    for day in week:
      if day == 0:
        full_html_doc += '<td class="cal-td cal-empty"></td>'
      else:
        is_today = (
            selected_year == today.year
            and selected_month_num == today.month
            and day == today.day
        )
        cell_class = "cal-td cal-today" if is_today else "cal-td"

        full_html_doc += f'<td class="{cell_class}"><div class="cal-day-num">{day}</div>'

        if day in daily_exercises:
          for ex in daily_exercises[day]:
            css_class = ex_class_map.get(ex, "ex-color-0")
            full_html_doc += f'<div class="cal-badge {css_class}" title="{ex}">{ex}</div>'

        full_html_doc += "</td>"
    full_html_doc += "</tr>"

  full_html_doc += """
        </tbody>
    </table>
    </body>
    </html>
    """

  # Render via iframe component to prevent markdown escaping
  components.html(full_html_doc, height=620, scrolling=True)

  st.markdown("---")

  # -------------------------------------------------------------------------
  # 2. PIE CHART
  # -------------------------------------------------------------------------
  st.subheader("💪 Set Distribution by Muscle Group")
  mg_counts = df_filtered["Grup Muscular"].value_counts().reset_index()
  mg_counts.columns = ["Grup Muscular", "Sets"]

  fig_pie = px.pie(
      mg_counts,
      values="Sets",
      names="Grup Muscular",
      color_discrete_sequence=px.colors.qualitative.Set2,
      hole=0.4,
  )
  fig_pie.update_layout(height=400)
  st.plotly_chart(fig_pie, use_container_width=True)

  st.markdown("---")

  # -------------------------------------------------------------------------
  # 3. RAW DATA LOG
  # -------------------------------------------------------------------------
  st.subheader("📋 Filtered Raw Data Log")
  st.dataframe(
      df_filtered[[
          "Data",
          "Exercici",
          "Grup Muscular",
          "Pes (kg)",
          "Repeticions",
          "Temps (min)",
          "Set_Volume",
          "Estimated_1RM",
      ]].sort_values(by="Data", ascending=False),
      height=400,
      use_container_width=True,
      hide_index=True,
  )